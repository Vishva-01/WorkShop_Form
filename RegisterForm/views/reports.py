from django.http import HttpResponse
from RegisterForm.models import Students
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle
from openpyxl.drawing.image import Image
import datetime

def export_students_to_excel(request):
    students = Students.objects.all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Students'

    headers = [
        'Name', 'College Name', 'Phone Number', 'WhatsApp Number', 'Email',
        'Gender', 'DOB', 'Department', 'Year Of Passing', 'Is Paid', 'Enrolled On'
    ]

    # Append the header to the sheet
    sheet.append(headers)

    for student in students:
        row = [
            student.name,
            student.college_name,
            student.phone_number,
            student.whatsapp_number,
            student.email,
            student.get_gender_display(),
            student.dob.strftime("%Y-%m-%d"),
            student.department,
            'Before 2022' if student.year==2021 else student.year,
            'Yes' if student.is_paid else 'No',
            student.enrolled_on.strftime("%Y-%m-%d/%H:%M")
        ]
        sheet.append(row)

    
    for col in range(1, len(headers) + 1):
        max_length = 0
        column = get_column_letter(col)
        for cell in sheet[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=students_report.xlsx'

    workbook.save(response)

    return response
